from django.contrib import admin
from django.template.defaultfilters import truncatechars
from simple_history.admin import SimpleHistoryAdmin
from .models import Category, UserCategoryPreference, Question, QuestionRating, Answer, AnswerRating, Bookmark
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from .inlines import AnswerInline


def export_to_excel(modeladmin, request, queryset):
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = 'Questions'

	columns = ['ID', 'Title', 'Author', 'Publication Date', 'Rating']
	for col_num, column_title in enumerate(columns, 1):
		cell = ws.cell(row=1, column=col_num)
		cell.value = column_title

	for row_num, question in enumerate(queryset, 2):
		ws.cell(row=row_num, column=1).value = question.pk
		ws.cell(row=row_num, column=2).value = question.title
		ws.cell(row=row_num, column=3).value = question.author.username
		ws.cell(row=row_num, column=4).value = question.pub_date.strftime('%Y-%m-%d %H:%M:%S')
		ws.cell(row=row_num, column=5).value = question.get_rating()

	for col_num in range(1, len(columns) + 1):
		column_letter = get_column_letter(col_num)
		ws.column_dimensions[column_letter].auto_size = True

	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = 'attachment; filename="questions.xlsx"'
	wb.save(response)
	return response


# Регистрируем действие
export_to_excel.short_description = "Export Selected to Excel"


# Register your models here.

class QuestionAdmin(SimpleHistoryAdmin):
	list_display = ('title_short', 'author', 'pub_date', 'get_rating', 'is_closed')
	list_filter = ('author', 'pub_date', 'categories', 'is_closed')
	search_fields = ('title', 'author__username', 'text')
	date_hierarchy = 'pub_date'
	inlines = [AnswerInline, ]
	readonly_fields = ('get_rating', 'pub_date', 'author')
	actions = ['make_closed', 'export_to_excel']
	fieldsets = (
		(None, {
			'fields': ('title', 'text', 'categories')
		}),
		('Advanced options', {
			'classes': ('collapse',),
			'fields': ('is_closed',),
		}),
	)

	def make_closed(self, request, queryset):
		queryset.update(is_closed=True)

	make_closed.short_description = "Mark selected questions as closed"


class QuestionAdmin(admin.ModelAdmin):
	list_display = ('title_short', 'author', 'pub_date', 'get_rating')
	list_filter = ('author', 'pub_date', 'categories')
	search_fields = ('title', 'author__username')
	date_hierarchy = 'pub_date'
	inlines = [AnswerInline, ]
	readonly_fields = ('get_rating',)
	filter_horizontal = ('categories',)
	actions = [export_to_excel]

	def get_rating(self, obj):
		return obj.get_rating()

	def title_short(self, obj):
		return truncatechars(obj.title, 10)

	title_short.short_description = 'Title'
	get_rating.short_description = 'Rating'


class AnswerAdmin(admin.ModelAdmin):
	list_display = ('text_short', 'question_short', 'user', 'pub_date', 'get_rating')
	list_filter = ('user', 'pub_date')
	search_fields = ('question__title', 'user__username', 'text')
	raw_id_fields = ('question',)

	def get_rating(self, obj):
		return obj.get_rating()

	def text_short(self, obj):
		return truncatechars(obj.text, 10)

	def question_short(self, obj):
		return truncatechars(obj.question.title, 10)

	text_short.short_description = 'Text'
	get_rating.short_description = 'Rating'


class QuestionRatingAdmin(admin.ModelAdmin):
	list_display = ('question_short', 'user', 'rating')
	list_filter = ('user', 'rating')
	raw_id_fields = ('question',)

	def question_short(self, obj):
		return truncatechars(obj.question.title, 10)


class AnswerRatingAdmin(admin.ModelAdmin):
	list_display = ('answer_short', 'user', 'rating')
	list_filter = ('user', 'rating')
	raw_id_fields = ('answer',)

	def answer_short(self, obj):
		return truncatechars(obj.answer.text, 10)


class BookmarkAdmin(admin.ModelAdmin):
	list_display = ('user', 'question_short', 'created_at')
	list_filter = ('user', 'created_at')
	raw_id_fields = ('question',)

	def question_short(self, obj):
		return truncatechars(obj.question.title, 10)


admin.site.register(Category)
admin.site.register(UserCategoryPreference)
admin.site.register(Question, QuestionAdmin)
admin.site.register(QuestionRating, QuestionRatingAdmin)
admin.site.register(Answer, AnswerAdmin)
admin.site.register(AnswerRating, AnswerRatingAdmin)
admin.site.register(Bookmark, BookmarkAdmin)
